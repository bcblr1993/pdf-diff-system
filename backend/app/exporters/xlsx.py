"""Excel 审核报告导出。

三个工作表：
- 摘要：任务信息 + 分类统计
- 差异清单：所有差异条目（含审核结论 + 批注 + 审核人）
- 关键差异：仅 critical 级条目，置顶展示给法务/商务
"""
from __future__ import annotations
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from sqlalchemy import select
from app.db.models import Comparison, Diff, DiffCategory, DiffSeverity, ReviewAction, User


CATEGORY_LABEL = {
    "replace": "修改",
    "delete": "删除",
    "insert": "新增",
    "handwritten": "手写填空",
    "stamp_covered": "章遮挡",
    "moved": "位置移动",
}

SEVERITY_LABEL = {"critical": "关键", "normal": "普通", "info": "信息"}
REVIEW_LABEL = {"confirmed": "✓ 确认", "ignored": "✗ 忽略", None: "—"}

CATEGORY_FILL = {
    "replace": PatternFill("solid", fgColor="FEF3C7"),
    "delete": PatternFill("solid", fgColor="FEE2E2"),
    "insert": PatternFill("solid", fgColor="DCFCE7"),
    "handwritten": PatternFill("solid", fgColor="DCFCE7"),
    "stamp_covered": PatternFill("solid", fgColor="E5E7EB"),
    "moved": PatternFill("solid", fgColor="DBEAFE"),
}

THIN_BORDER = Border(
    left=Side(border_style="thin", color="D1D5DB"),
    right=Side(border_style="thin", color="D1D5DB"),
    top=Side(border_style="thin", color="D1D5DB"),
    bottom=Side(border_style="thin", color="D1D5DB"),
)


def export_xlsx(db: Session, comparison_id: int) -> bytes:
    cmp = db.get(Comparison, comparison_id)
    if not cmp:
        raise ValueError("Comparison 不存在")

    diffs = db.scalars(
        select(Diff).where(Diff.comparison_id == comparison_id).order_by(Diff.seq_no.asc())
    ).all()

    # 收集涉及的审核人 displayname
    reviewer_ids = {d.reviewed_by for d in diffs if d.reviewed_by} | (
        {cmp.review_completed_by} if cmp.review_completed_by else set()
    )
    reviewer_map = {}
    if reviewer_ids:
        users = db.scalars(select(User).where(User.id.in_(reviewer_ids))).all()
        reviewer_map = {u.id: u.display_name or u.username for u in users}

    wb = Workbook()
    _build_summary_sheet(wb, cmp, diffs, reviewer_map)
    _build_diffs_sheet(wb.create_sheet("差异清单"), cmp, diffs, reviewer_map, only_real=False)
    _build_diffs_sheet(
        wb.create_sheet("关键差异"), cmp,
        [d for d in diffs if d.severity == DiffSeverity.critical],
        reviewer_map, only_real=True,
    )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_summary_sheet(wb: Workbook, cmp: Comparison, diffs: list[Diff], reviewer_map: dict):
    ws = wb.active
    ws.title = "摘要"

    title_font = Font(bold=True, size=14, color="1F2937")
    label_font = Font(bold=True, color="6B7280")

    ws["A1"] = "PDF 差异对比审核报告"
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells("A1:D1")

    rows = [
        ("任务标题", cmp.title),
        ("任务编号", f"#{cmp.id}"),
        ("状态", cmp.status.value),
        ("审核状态", cmp.review_status.value),
        ("原件文件", cmp.orig_file.original_name if cmp.orig_file else "—"),
        ("扫描件文件", cmp.scan_file.original_name if cmp.scan_file else "—"),
        ("创建时间", _fmt_dt(cmp.created_at)),
        ("完成时间", _fmt_dt(cmp.completed_at)),
        ("审核完成时间", _fmt_dt(cmp.review_completed_at)),
        ("审核完成人", reviewer_map.get(cmp.review_completed_by, "—") if cmp.review_completed_by else "—"),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=k).font = label_font
        ws.cell(row=i, column=2, value=str(v))

    # 统计区
    s = cmp.summary_json or {}
    start = 3 + len(rows) + 2
    ws.cell(row=start, column=1, value="差异统计").font = title_font

    counts = [
        ("真实差异（总）", s.get("real", 0), "FED7AA"),
        ("★ 关键字段", s.get("critical", 0), "FCA5A5"),
        ("修改", s.get("replace", 0), "FEF3C7"),
        ("删除", s.get("delete", 0), "FEE2E2"),
        ("新增", s.get("insert", 0) + s.get("handwritten", 0), "DCFCE7"),
        ("章遮挡", s.get("stamp_covered", 0), "E5E7EB"),
        ("位置移动（噪声）", s.get("moved", 0), "DBEAFE"),
        ("页眉页脚（噪声）", s.get("footer", 0), "F3F4F6"),
        ("总条目（含噪声）", s.get("total", 0), "F9FAFB"),
    ]
    for i, (label, value, color) in enumerate(counts, start=start + 1):
        ws.cell(row=i, column=1, value=label).font = label_font
        cell = ws.cell(row=i, column=2, value=value)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center")

    # 审核进度
    reviewed = sum(1 for d in diffs if d.review_action is not None)
    confirmed = sum(1 for d in diffs if d.review_action == ReviewAction.confirmed)
    ignored = sum(1 for d in diffs if d.review_action == ReviewAction.ignored)
    review_start = start + len(counts) + 2
    ws.cell(row=review_start, column=1, value="审核进度").font = title_font
    review_rows = [
        ("已审核", f"{reviewed} / {len(diffs)}"),
        ("✓ 确认", confirmed),
        ("✗ 忽略", ignored),
        ("未审核", len(diffs) - reviewed),
    ]
    for i, (k, v) in enumerate(review_rows, start=review_start + 1):
        ws.cell(row=i, column=1, value=k).font = label_font
        ws.cell(row=i, column=2, value=v)

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 60
    for col in ("C", "D"):
        ws.column_dimensions[col].width = 14


def _build_diffs_sheet(ws, cmp: Comparison, diffs: list[Diff], reviewer_map: dict, *, only_real: bool):
    headers = ["#", "类别", "严重度", "原件页", "扫描件页", "原件文本", "扫描件文本",
               "上下文", "审核结论", "批注", "审核人", "审核时间"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="374151")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    NOISE_CATEGORIES = {DiffCategory.moved}

    row_idx = 2
    for d in diffs:
        # only_real 模式下跳过噪声
        if only_real and (
            d.category in NOISE_CATEGORIES
            or d.severity == DiffSeverity.info
            or d.is_footer
        ):
            continue
        vals = [
            d.seq_no,
            CATEGORY_LABEL.get(d.category.value, d.category.value),
            ("★ " if d.severity == DiffSeverity.critical else "") + SEVERITY_LABEL.get(d.severity.value, ""),
            d.orig_page + 1 if d.orig_page >= 0 else "—",
            d.scan_page + 1 if d.scan_page >= 0 else "—",
            d.orig_text or "",
            d.scan_text or "",
            d.context or "",
            REVIEW_LABEL.get(d.review_action.value if d.review_action else None, "—"),
            d.review_note or "",
            reviewer_map.get(d.reviewed_by, "—") if d.reviewed_by else "—",
            _fmt_dt(d.reviewed_at),
        ]
        for col, v in enumerate(vals, start=1):
            cell = ws.cell(row=row_idx, column=col, value=v)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col == 2:
                cell.fill = CATEGORY_FILL.get(d.category.value, PatternFill())
            if d.severity == DiffSeverity.critical and col == 3:
                cell.font = Font(bold=True, color="DC2626")
            if d.review_action == ReviewAction.confirmed and col == 9:
                cell.font = Font(bold=True, color="DC2626")
            elif d.review_action == ReviewAction.ignored and col == 9:
                cell.font = Font(color="6B7280")
        row_idx += 1

    widths = [6, 12, 10, 8, 8, 36, 36, 40, 12, 30, 10, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def _fmt_dt(dt) -> str:
    if not dt:
        return "—"
    if isinstance(dt, str):
        return dt
    return dt.strftime("%Y-%m-%d %H:%M:%S") if isinstance(dt, datetime) else str(dt)
