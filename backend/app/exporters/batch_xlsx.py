"""批量任务汇总 Excel：总览 sheet + 每份扫描件一个 sheet。"""
from __future__ import annotations
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session
from app.db.models import BatchJob, Comparison, Diff, DiffSeverity, ReviewAction, User
from app.exporters.xlsx import (
    CATEGORY_LABEL, REVIEW_LABEL, _build_diffs_sheet, _fmt_dt, THIN_BORDER,
)


def export_batch_xlsx(db: Session, batch_id: int) -> bytes:
    batch = db.get(BatchJob, batch_id)
    if not batch:
        raise ValueError("批量任务不存在")

    comps = db.scalars(
        select(Comparison).where(Comparison.batch_id == batch_id).order_by(Comparison.id.asc())
    ).all()

    # 收集所有审核人 displayname
    all_reviewer_ids: set[int] = set()
    diffs_by_cmp: dict[int, list[Diff]] = {}
    for cmp in comps:
        diffs = db.scalars(
            select(Diff).where(Diff.comparison_id == cmp.id).order_by(Diff.seq_no.asc())
        ).all()
        diffs_by_cmp[cmp.id] = diffs
        for d in diffs:
            if d.reviewed_by:
                all_reviewer_ids.add(d.reviewed_by)
        if cmp.review_completed_by:
            all_reviewer_ids.add(cmp.review_completed_by)
    reviewer_map = {}
    if all_reviewer_ids:
        users = db.scalars(select(User).where(User.id.in_(all_reviewer_ids))).all()
        reviewer_map = {u.id: u.display_name or u.username for u in users}

    wb = Workbook()
    _build_overview(wb, batch, comps, diffs_by_cmp, reviewer_map)

    # 每份扫描件一个 sheet
    used_names: set[str] = {"总览"}
    for cmp in comps:
        name = _safe_sheet_name(cmp.scan_file.original_name if cmp.scan_file else f"对比-{cmp.id}", used_names)
        used_names.add(name)
        ws = wb.create_sheet(name)
        _build_diffs_sheet(ws, cmp, diffs_by_cmp.get(cmp.id, []), reviewer_map, only_real=False)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_overview(wb: Workbook, batch: BatchJob, comps: list[Comparison],
                    diffs_by_cmp: dict, reviewer_map: dict):
    ws = wb.active
    ws.title = "总览"

    title_font = Font(bold=True, size=16)
    label_font = Font(bold=True, color="6B7280")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="374151")

    # ── 标题与元信息 ──
    ws["A1"] = f"批量对比汇总 - {batch.title or f'#{batch.id}'}"
    ws["A1"].font = title_font
    ws.merge_cells("A1:H1")

    meta_rows = [
        ("批量编号", f"#{batch.id}"),
        ("原件文件", batch.orig_file.original_name if batch.orig_file else "—"),
        ("扫描件数量", str(batch.total)),
        ("已完成", f"{batch.completed} / {batch.total}"),
        ("失败", str(batch.failed)),
        ("状态", batch.status.value),
        ("创建时间", _fmt_dt(batch.created_at)),
        ("完成时间", _fmt_dt(batch.completed_at)),
    ]
    for i, (k, v) in enumerate(meta_rows, start=3):
        ws.cell(row=i, column=1, value=k).font = label_font
        ws.cell(row=i, column=2, value=v)

    # ── 子任务一览表 ──
    start = 3 + len(meta_rows) + 2
    ws.cell(row=start, column=1, value="子任务一览").font = Font(bold=True, size=12)

    headers = ["#", "扫描件文件", "状态", "真实差异", "关键★", "新增", "删除", "修改",
               "已审核", "审核完成"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=start + 1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    for i, cmp in enumerate(comps, start=start + 2):
        s = cmp.summary_json or {}
        diffs = diffs_by_cmp.get(cmp.id, [])
        reviewed = sum(1 for d in diffs if d.review_action)
        # 排除噪声计算真实差异审核进度
        real_diffs = [d for d in diffs if d.category.value != "moved"
                      and d.severity.value != "info" and not d.is_footer]
        real_reviewed = sum(1 for d in real_diffs if d.review_action)

        vals = [
            cmp.id,
            cmp.scan_file.original_name if cmp.scan_file else "—",
            cmp.status.value,
            s.get("real", 0),
            s.get("critical", 0),
            s.get("insert", 0) + s.get("handwritten", 0),
            s.get("delete", 0),
            s.get("replace", 0),
            f"{real_reviewed} / {len(real_diffs)}",
            reviewer_map.get(cmp.review_completed_by, "—") if cmp.review_completed_by else "—",
        ]
        for col, v in enumerate(vals, start=1):
            cell = ws.cell(row=i, column=col, value=v)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col == 5 and isinstance(v, int) and v > 0:
                cell.font = Font(bold=True, color="DC2626")
                cell.fill = PatternFill("solid", fgColor="FEF2F2")
            if col == 3:
                if v == "failed":
                    cell.fill = PatternFill("solid", fgColor="FEE2E2")
                elif v == "done":
                    cell.fill = PatternFill("solid", fgColor="DCFCE7")

    widths = [6, 40, 10, 10, 8, 8, 8, 8, 12, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{start + 2}"


def _safe_sheet_name(name: str, used: set[str]) -> str:
    """Excel sheet 名限制：≤31 字符，不能含 / \\ ? * [ ] : """
    name = name.replace("/", "_").replace("\\", "_").replace("?", "")
    name = name.replace("*", "").replace("[", "(").replace("]", ")").replace(":", "_")
    name = name[:31]
    base = name
    i = 1
    while name in used:
        suffix = f"_{i}"
        name = base[: 31 - len(suffix)] + suffix
        i += 1
    return name
